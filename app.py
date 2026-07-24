import streamlit as st
import pandas as pd
import json
from io import BytesIO
import pytz
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Step Cadence Analyzer", layout="wide")
st.title("Minute-by-Minute Step Cadence Analyzer")

# --- INSTRUCTIONS SECTION ---
with st.expander("📋 How to Use This Tool", expanded=True):
    st.markdown("""
    This application processes raw, minute-by-minute step data exported from Fitbit devices 
    and automatically calculates daily physical activity summaries based on step cadence bands.
    It also produces a complete minute-level log, hourly activity tables, and an exploratory
    red-flag summary designed to help researcher identify days that may warrant review for incomplete
    data coverage.
    
    ---
    
    **Step 1: Export Your Fitbit Data**
    
    Visit the [Fitbit Data Export Help Page](https://support.google.com/googlehealth/answer/14236615?hl=en) 
    for step-by-step instructions on downloading your Fitbit data archive. Note that instructions 
    may vary slightly depending on whether the participant signs in with a Google Account or a 
    Fitbit login.
    
    Once downloaded, unzip the archive. You will need:
    - The files named **"steps-YYYY-MM-DD.json"** (e.g., steps-2012-10-01.json) 
      from the **"Physical Activity"** folder
    - Optionally, **"Profile.csv"** from the **"Personal & Account"** folder so the
      tool can identify the participant's Fitbit profile timezone
    
    ---
    
    **Step 2: Configure Settings**
    
    Before processing the step data, you must confirm the participant's timezone. In the
    sidebar under **A. Timezone**, either:
    - Upload **"Profile.csv"** under **A** and allow the tool to select the timezone recorded in the
      participant's Fitbit profile; or
    - Manually select the **timezone** that matches the participant's location during data
      collection if **"Profile.csv"** is unavailable

    Review the displayed timezone and check the confirmation box before continuing. Also use
    the sidebar to:
    - Enter a unique **Participant ID** under **B** (this label will appear in your output files)
    - Adjust the **intensity thresholds** under **C** if your study uses different MPA and
      VPA criteria (defaults: MPA ≥100 spm and VPA ≥130 spm)²
    - Adjust the **cadence band thresholds** under **D** if your study uses custom step-rate criteria
      (default values and names follow Tudor-Locke et al.¹)

    Fitbit minute-level step timestamps are interpreted as UTC and converted to the confirmed
    participant timezone before daily and hourly summaries are calculated. This conversion
    automatically applies the correct daylight-saving offset for each date.
    
    ---
    
    **Step 3: Upload Files and Select the Analysis Dates**
    
    Upload all JSON step files for the participant using **E. Data Upload** in the sidebar.
    After the files are read and converted to local time, the app will display the earliest
    and latest available dates. Select the intended analysis start and end dates, then confirm
    the range. The same selected date range will be applied to all four output files.

    The app will then process the selected dates. Download your results using the buttons
    that appear when processing is complete.
    
    Four output files are provided:
    - 📄 **Daily Summary (CSV):** Total steps; minutes in each of the cadence bands; and MPA, VPA, and total MVPA minutes per calendar day
    - 📄 **Minute-by-Minute Log (CSV):** A complete chronological record with both the cadence-band and intensity classification for each minute
    - 📊 **Hourly Analysis (CVS):** Total steps and minutes per cadence band broken down by hour of day, with separate MPA and VPA tabs
    - 🚩 **Red Flag Summary (CVS):** Daily and daypart steps, active minutes,
      plausible wear minutes, activity span, and four exploratory data-coverage
      indicators in a single worksheet
    
    ---
    
    **Step 4: Processing a New Participant**
    
    To prevent data from one participant from carrying over to the next, 
    **refresh your browser before uploading files for a new participant.** 
    This clears the session and fully resets all settings.
    
    ---
    
    ⚠️ *This tool does not store, save, or transmit any uploaded data. All files are 
    processed temporarily and deleted automatically when the browser is refreshed or closed.*

    ---

    **References for Default Settings**

    ¹ Tudor-Locke, C., Camhi, S. M., Leonardi, C., Johnson, W. D.,
    Katzmarzyk, P. T., Earnest, C. P., & Church, T. S. (2011). Patterns of adult
    stepping cadence in the 2005–2006 NHANES. *Preventive Medicine, 53*(3),
    178–181. https://doi.org/10.1016/j.ypmed.2011.06.004

    ² O'Brien, M. W., Kivell, M. J., Wojcik, W. R., D'Entremont, G.,
    Kimmerly, D. S., & Fowles, J. R. (2018). Step rate thresholds associated
    with moderate and vigorous physical activity in adults. *International
    Journal of Environmental Research and Public Health, 15*(11), 2454.
    https://doi.org/10.3390/ijerph15112454
    """)

# --- SIDEBAR ---
st.sidebar.header("A. Timezone")
profile_file = st.sidebar.file_uploader(
    'Optional: Upload "Profile.csv"',
    type=['csv'],
    accept_multiple_files=False,
    help='In the unzipped Fitbit export, find Profile.csv in the "Personal & Account" folder.'
)

detected_timezone = None
if profile_file is not None:
    try:
        profile_df = pd.read_csv(profile_file)
        profile_file.seek(0)

        if 'timezone' not in profile_df.columns or profile_df.empty:
            st.sidebar.warning(
                '⚠️ The uploaded CSV does not contain a usable "timezone" field. '
                'Please select the timezone manually.'
            )
        else:
            candidate_timezone = str(profile_df.loc[0, 'timezone']).strip()
            if candidate_timezone in pytz.all_timezones:
                detected_timezone = candidate_timezone
                st.sidebar.success(
                    f"✅ Timezone found in Profile.csv: {detected_timezone}"
                )
            else:
                st.sidebar.warning(
                    f'⚠️ The timezone in Profile.csv ("{candidate_timezone}") was not '
                    'recognized. Please select the timezone manually.'
                )
    except Exception as e:
        st.sidebar.warning(
            f"⚠️ Profile.csv could not be read ({e}). Please select the timezone manually."
        )

timezone_options = [
    "America/New_York",
    "America/Chicago",
    "America/Denver",
    "America/Los_Angeles",
    "UTC"
]
if detected_timezone and detected_timezone not in timezone_options:
    timezone_options.insert(0, detected_timezone)

default_timezone = detected_timezone or "America/New_York"
timezone = st.sidebar.selectbox(
    "Confirm or manually select timezone",
    timezone_options,
    index=timezone_options.index(default_timezone),
    help="This should match the participant's location during data collection."
)
timezone_confirmed = st.sidebar.checkbox(
    f"I confirm the participant timezone is {timezone}",
    value=False
)

st.sidebar.header("B. Participant Information")
manual_participant_id = st.sidebar.text_input("Enter Participant ID", value="Participant_1")

st.sidebar.header("C. Intensity Thresholds (spm)")
st.sidebar.markdown(
    "Default intensity thresholds are MPA ≥100 spm and VPA ≥130 spm "
    "(O'Brien et al., 2018)."
)
mpa = st.sidebar.number_input("MPA (Moderate Physical Activity) lower limit", value=100)
vpa = st.sidebar.number_input("VPA (Vigorous Physical Activity) lower limit", value=130)

st.sidebar.header("D. Cadence Band Thresholds (spm)")
st.sidebar.markdown(
    "Default band names and thresholds follow Tudor-Locke et al. (2011)."
)
st.sidebar.markdown("---")
b1 = st.sidebar.number_input("Incidental movement lower limit", value=1)
b2 = st.sidebar.number_input("Sporadic movement lower limit", value=20)
b3 = st.sidebar.number_input("Purposeful steps lower limit", value=40)
b4 = st.sidebar.number_input("Slow walking lower limit", value=60)
b5 = st.sidebar.number_input("Medium walking lower limit", value=80)
b6 = st.sidebar.number_input("Brisk walking lower limit", value=100)
b7 = st.sidebar.number_input("Faster locomotion/ambulation lower limit", value=120)

st.sidebar.header("E. Data Upload")
uploaded_files = st.sidebar.file_uploader(
    "Upload Fitbit JSON files",
    type=['json'],
    accept_multiple_files=True
)

# Hour labels for Excel output (hour 0 = midnight to 1am)
hour_labels = [
    "12:00-1:00 AM", "1:00-2:00 AM", "2:00-3:00 AM", "3:00-4:00 AM",
    "4:00-5:00 AM", "5:00-6:00 AM", "6:00-7:00 AM", "7:00-8:00 AM",
    "8:00-9:00 AM", "9:00-10:00 AM", "10:00-11:00 AM", "11:00 AM-12:00 PM",
    "12:00-1:00 PM", "1:00-2:00 PM", "2:00-3:00 PM", "3:00-4:00 PM",
    "4:00-5:00 PM", "5:00-6:00 PM", "6:00-7:00 PM", "7:00-8:00 PM",
    "8:00-9:00 PM", "9:00-10:00 PM", "10:00-11:00 PM", "11:00 PM-12:00 AM"
]


DAYPARTS = ["Overnight", "Morning", "Afternoon", "Evening"]


def format_clock(timestamp):
    """Format a timestamp as, for example, 6:33 p.m."""
    hour = timestamp.hour % 12 or 12
    suffix = "a.m." if timestamp.hour < 12 else "p.m."
    return f"{hour}:{timestamp.minute:02d} {suffix}"


def build_red_flag_workbook(part_df, participant_id):
    """Build the single-sheet exploratory daily/daypart data-coverage summary."""
    red_flag_df = part_df[['Date', 'Steps']].copy()
    red_flag_df['Daypart'] = pd.cut(
        red_flag_df.index.hour,
        bins=[-1, 5, 11, 17, 23],
        labels=DAYPARTS
    )
    red_flag_df['Active_Minute'] = red_flag_df['Steps'].gt(0).astype(int)

    # Modified step-based nonwear proxy: every step-positive minute interrupts
    # a run. Only uninterrupted zero-step runs of >=90 consecutive minutes are
    # classified as prolonged zero-step time.
    zero_step = red_flag_df['Steps'].eq(0)
    consecutive_minute = red_flag_df.index.to_series().diff().eq(pd.Timedelta(minutes=1))
    run_start = (~zero_step) | (~consecutive_minute)
    run_id = run_start.cumsum()
    zero_run_lengths = zero_step.groupby(run_id).transform('sum')
    red_flag_df['Prolonged_Zero_Step'] = (
        zero_step & zero_run_lengths.ge(90)
    ).astype(int)
    red_flag_df['Plausible_Wear_Minute'] = 1 - red_flag_df['Prolonged_Zero_Step']

    dates = sorted(red_flag_df['Date'].unique())
    rows = []
    for day_number, date_value in enumerate(dates, start=1):
        day_df = red_flag_df[red_flag_df['Date'] == date_value]
        steps_by_part = (
            day_df.groupby('Daypart', observed=False)['Steps']
            .sum().reindex(DAYPARTS, fill_value=0).astype(int)
        )
        active_by_part = (
            day_df.groupby('Daypart', observed=False)['Active_Minute']
            .sum().reindex(DAYPARTS, fill_value=0).astype(int)
        )
        wear_by_part = (
            day_df.groupby('Daypart', observed=False)['Plausible_Wear_Minute']
            .sum().reindex(DAYPARTS, fill_value=0).astype(int)
        )
        active_rows = day_df[day_df['Steps'] > 0]
        if active_rows.empty:
            activity_span_label = ""
            activity_span_minutes = None
        else:
            first_active = active_rows.index[0]
            last_active = active_rows.index[-1]
            activity_span_label = (
                f"{format_clock(first_active)}–{format_clock(last_active)}"
            )
            activity_span_minutes = int(
                (last_active - first_active).total_seconds() // 60
            )

        total_steps = int(steps_by_part.sum())
        total_active = int(active_by_part.sum())
        total_wear = int(wear_by_part.sum())
        active_dayparts = int((active_by_part > 0).sum())

        rows.append([
            day_number,
            total_steps, *steps_by_part.tolist(),
            total_active, *active_by_part.tolist(),
            total_wear, *wear_by_part.tolist(),
            activity_span_label, activity_span_minutes,
            "RED FLAG" if total_steps < 1000 else "",
            "RED FLAG" if active_dayparts == 1 else "",
            (
                "RED FLAG"
                if activity_span_minutes is not None
                and activity_span_minutes < 360
                else ""
            ),
            "RED FLAG" if total_wear < 600 else "",
        ])
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        summary_name = f"{participant_id} Red Flag Summary"[:31]
        summary_sheet = workbook.create_sheet(summary_name)
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        group_headers = [
            "Day",
            "Steps", None, None, None, None,
            "Active Minutes (>0 Steps)", None, None, None, None,
            "Plausible Wear Minutes", None, None, None, None,
            "Activity Span", None,
            "Red-Flag Indicators", None, None, None,
        ]
        subheaders = [
            None,
            "Total", *DAYPARTS,
            "Total", *DAYPARTS,
            "Total", *DAYPARTS,
            "First–Last Step-Positive Time", "Span (Minutes)",
            "Total Steps <1,000",
            "Activity in Only One Daypart",
            "Activity Span <6 Hours",
            "Plausible Wear <600 Minutes",
        ]
        summary_sheet.append(group_headers)
        summary_sheet.append(subheaders)
        for row in rows:
            summary_sheet.append(row)

        summary_sheet.merge_cells("A1:A2")
        for cell_range in ["B1:F1", "G1:K1", "L1:P1", "Q1:R1", "S1:V1"]:
            summary_sheet.merge_cells(cell_range)

        header_colors = {
            "A": "274C77", "B": "4F81BD", "G": "548A54",
            "L": "C97A28", "Q": "695D8F", "S": "A61B1B",
        }
        for start_col, end_col in [(1, 1), (2, 6), (7, 11), (12, 16), (17, 18), (19, 22)]:
            fill = PatternFill("solid", fgColor=header_colors[get_column_letter(start_col)])
            for col in range(start_col, end_col + 1):
                cell = summary_sheet.cell(1, col)
                cell.fill = fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        subheader_colors = [
            (2, 6, "D9EAF7"), (7, 11, "DDEEDB"), (12, 16, "F9E2C7"),
            (17, 18, "E4DFF2"), (19, 22, "F4CCCC"),
        ]
        for start_col, end_col, color in subheader_colors:
            for col in range(start_col, end_col + 1):
                cell = summary_sheet.cell(2, col)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="243240")
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        thin_gray = Side(style="thin", color="C7D0D9")
        medium_gray = Side(style="medium", color="9AA7B2")
        for row in summary_sheet.iter_rows(
            min_row=1, max_row=summary_sheet.max_row, min_col=1, max_col=22
        ):
            for cell in row:
                cell.border = Border(bottom=thin_gray)
                if cell.column in [6, 11, 16, 18, 22]:
                    cell.border = Border(bottom=thin_gray, right=medium_gray)
                if cell.row >= 3 and cell.column >= 19:
                    cell.font = Font(bold=True, color="A61B1B")
                    cell.alignment = Alignment(horizontal="center")

        widths = {
            1: 8, 17: 24, 18: 14, 19: 18, 20: 20, 21: 18, 22: 20,
        }
        for col in range(2, 17):
            widths.setdefault(col, 12)
        for col, width in widths.items():
            summary_sheet.column_dimensions[get_column_letter(col)].width = width
        summary_sheet.row_dimensions[1].height = 28
        summary_sheet.row_dimensions[2].height = 42
        summary_sheet.freeze_panes = "B3"
        summary_sheet.sheet_view.showGridLines = False

    output.seek(0)
    return output


if uploaded_files and not timezone_confirmed:
    st.warning(
        "⚠️ Confirm the participant's timezone in the sidebar before processing the step files."
    )

elif uploaded_files:
    st.write(f"### Reading {len(uploaded_files)} file(s) for {manual_participant_id}...")

    all_raw_dfs = []

    # 1. READ ALL FILES WITH SCHEMA VALIDATION
    for file in uploaded_files:
        try:
            data = json.load(file)
        except Exception:
            st.warning(f"⚠️ Skipping `{file.name}`: Not a valid JSON file.")
            continue

        if not data or not isinstance(data, list):
            st.warning(f"⚠️ Skipping `{file.name}`: File is empty or not in the expected Fitbit format.")
            continue

        temp_df = pd.DataFrame(data)

        if 'dateTime' not in temp_df.columns or 'value' not in temp_df.columns:
            st.warning(f"⚠️ Skipping `{file.name}`: Missing 'dateTime' or 'value' columns. Are you sure this is a steps file?")
            continue

        try:
            temp_df['value'] = temp_df['value'].astype(int)
            # Fitbit account-export step timestamps are interpreted as UTC.
            # The UTC-aware values are converted to the participant's confirmed
            # timezone after all uploaded files are combined.
            temp_df['dateTime'] = pd.to_datetime(
                temp_df['dateTime'],
                errors='raise',
                utc=True
            )
        except Exception as e:
            st.warning(f"⚠️ Skipping `{file.name}`: Data formatting error. ({e})")
            continue

        all_raw_dfs.append(temp_df)

    if all_raw_dfs:
        df = pd.concat(all_raw_dfs)
        df = df.drop_duplicates(subset=['dateTime'])
        df = df.set_index('dateTime').tz_convert(timezone)
        df = df[df.index.notna()]

        earliest_available_date = df.index.min().date()
        latest_available_date = df.index.max().date()

        st.write("### Select Analysis Date Range")
        st.info(
            f"Available local dates: **{earliest_available_date:%B %d, %Y}** "
            f"through **{latest_available_date:%B %d, %Y}**"
        )

        date_col1, date_col2 = st.columns(2)
        with date_col1:
            analysis_start_date = st.date_input(
                "Analysis start date",
                value=earliest_available_date,
                min_value=earliest_available_date,
                max_value=latest_available_date
            )
        with date_col2:
            analysis_end_date = st.date_input(
                "Analysis end date",
                value=latest_available_date,
                min_value=earliest_available_date,
                max_value=latest_available_date
            )

        valid_date_range = analysis_start_date <= analysis_end_date
        if not valid_date_range:
            st.error("❌ The analysis start date must be on or before the end date.")

        date_range_confirmed = st.checkbox(
            (
                "I confirm that the analysis should include "
                f"{analysis_start_date:%B %d, %Y} through "
                f"{analysis_end_date:%B %d, %Y}"
            ),
            value=False,
            disabled=not valid_date_range
        )

        if valid_date_range:
            local_dates = pd.Series(df.index.date, index=df.index)
            selected_date_mask = (
                (local_dates >= analysis_start_date)
                & (local_dates <= analysis_end_date)
            )
            selected_df = df.loc[selected_date_mask.to_numpy()]
        else:
            selected_df = df.iloc[0:0]

        if valid_date_range and selected_df.empty:
            st.error("❌ No step records were found within the selected date range.")

        if not date_range_confirmed:
            st.info("Confirm the analysis date range to generate the four output files.")
            st.stop()

        if selected_df.empty:
            st.stop()

        df = selected_df
        st.write(
            f"### Processing {analysis_start_date:%B %d, %Y} through "
            f"{analysis_end_date:%B %d, %Y}..."
        )

        unique_dates = df.index.normalize().unique()
        participant_full_timeseries = []

        # 2. RECONSTRUCT CONTINUOUS TIMELINES PER DAY
        for date in unique_dates:
            if pd.isna(date):
                continue

            daily_index = pd.date_range(
                start=date,
                end=date + pd.DateOffset(days=1) - pd.Timedelta(minutes=1),
                freq='min',
                tz=timezone
            )

            day_data = df[df.index.normalize() == date]
            day_data = day_data[~day_data.index.duplicated(keep='first')]
            day_full = day_data.reindex(daily_index, fill_value=0)

            day_full['Participant_ID'] = manual_participant_id
            day_full['Date'] = date.date()
            day_full['Time'] = day_full.index.time
            day_full = day_full.rename(columns={'value': 'Steps'})

            participant_full_timeseries.append(day_full)

        part_df = pd.concat(participant_full_timeseries)

        # Add hour column for hourly Excel output
        part_df['Hour'] = part_df.index.hour

        # 3. DEFINE TUDOR-LOCKE CADENCE BINS AND INTENSITY CLASSIFICATIONS
        bins = [-1, b1-1, b2-1, b3-1, b4-1, b5-1, b6-1, b7-1, 9999]

        mpa_label = f'MPA ({mpa}-{vpa-1} spm)'
        vpa_label = f'VPA ({vpa}+ spm)'

        labels = [
            'Non-movement (0 spm)',
            f'Incidental movement ({b1}-{b2-1} spm)',
            f'Sporadic movement ({b2}-{b3-1} spm)',
            f'Purposeful steps ({b3}-{b4-1} spm)',
            f'Slow walking ({b4}-{b5-1} spm)',
            f'Medium walking ({b5}-{b6-1} spm)',
            f'Brisk walking ({b6}-{b7-1} spm)',
            f'Faster locomotion ({b7}+ spm)'
        ]

        part_df['Cadence_Band'] = pd.cut(part_df['Steps'], bins=bins, labels=labels)
        part_df['Intensity_Category'] = 'Below MPA threshold'
        part_df.loc[
            (part_df['Steps'] >= mpa) & (part_df['Steps'] < vpa),
            'Intensity_Category'
        ] = mpa_label
        part_df.loc[part_df['Steps'] >= vpa, 'Intensity_Category'] = vpa_label

        # 4. CALCULATE DAILY SUMMARIES
        summary = part_df.groupby(
            ['Participant_ID', 'Date', 'Cadence_Band'], observed=False
        ).size().unstack(fill_value=0)
        summary = summary.reindex(columns=labels, fill_value=0)

        daily_intensity = part_df.groupby(['Participant_ID', 'Date']).agg(
            Total_MPA_Minutes=(
                'Steps', lambda values: ((values >= mpa) & (values < vpa)).sum()
            ),
            Total_VPA_Minutes=('Steps', lambda values: (values >= vpa).sum())
        )
        summary['Total_MPA_Minutes'] = daily_intensity['Total_MPA_Minutes']
        summary['Total_VPA_Minutes'] = daily_intensity['Total_VPA_Minutes']
        summary['Total_MVPA_Minutes'] = summary['Total_MPA_Minutes'] + summary['Total_VPA_Minutes']
        summary['Total_Daily_Steps'] = part_df.groupby(['Participant_ID', 'Date'])['Steps'].sum()

        column_order = [
            'Total_Daily_Steps',
            'Total_MPA_Minutes',
            'Total_VPA_Minutes',
            'Total_MVPA_Minutes',
        ] + labels

        final_summary = summary[column_order].reset_index()
        final_summary = final_summary.sort_values(by='Date')

        final_min_by_min = part_df[
            [
                'Participant_ID',
                'Date',
                'Time',
                'Steps',
                'Cadence_Band',
                'Intensity_Category'
            ]
        ]

        # 5. BUILD HOURLY EXCEL OUTPUT
        all_dates = sorted(part_df['Date'].unique())

        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:

            analysis_information = pd.DataFrame({
                'Setting': [
                    'Participant ID',
                    'Available start date',
                    'Available end date',
                    'Selected analysis start date',
                    'Selected analysis end date',
                    'Participant timezone',
                    'Timestamp processing',
                    'Cadence-band defaults',
                    'Intensity defaults',
                    'Cadence-band reference',
                    'Intensity-threshold reference'
                ],
                'Value': [
                    manual_participant_id,
                    earliest_available_date.isoformat(),
                    latest_available_date.isoformat(),
                    analysis_start_date.isoformat(),
                    analysis_end_date.isoformat(),
                    timezone,
                    (
                        "Source timestamps interpreted as UTC and converted "
                        "to the participant timezone before aggregation"
                    ),
                    (
                        f"0; {b1}-{b2-1}; {b2}-{b3-1}; {b3}-{b4-1}; "
                        f"{b4}-{b5-1}; {b5}-{b6-1}; {b6}-{b7-1}; {b7}+ spm"
                    ),
                    f"MPA {mpa}-{vpa-1} spm; VPA {vpa}+ spm",
                    "Tudor-Locke et al. (2011), doi:10.1016/j.ypmed.2011.06.004",
                    "O'Brien et al. (2018), doi:10.3390/ijerph15112454"
                ]
            })
            analysis_information.to_excel(
                writer, sheet_name='Analysis Information', index=False
            )

            # Tab 1: Total Steps per Hour
            hourly_steps = part_df.groupby(['Date', 'Hour'])['Steps'].sum().unstack(fill_value=0)
            hourly_steps = hourly_steps.reindex(index=all_dates, fill_value=0)
            hourly_steps = hourly_steps.reindex(columns=range(24), fill_value=0)
            hourly_steps.columns = hour_labels
            hourly_steps.reset_index().to_excel(
                writer, sheet_name='Total Steps per Hour', index=False
            )

            # One tab per cadence band
            for label in labels:
                band_data = part_df[part_df['Cadence_Band'] == label]

                if not band_data.empty:
                    hourly_band = band_data.groupby(
                        ['Date', 'Hour']
                    ).size().unstack(fill_value=0)
                else:
                    hourly_band = pd.DataFrame(
                        index=pd.Index(all_dates, name='Date'),
                        columns=range(24),
                        dtype=float
                    ).fillna(0)

                hourly_band = hourly_band.reindex(index=all_dates, fill_value=0)
                hourly_band = hourly_band.reindex(columns=range(24), fill_value=0)
                hourly_band = hourly_band.astype(int)
                hourly_band.columns = hour_labels
                # Excel sheet names max 31 characters
                sheet_name = label[:31]
                hourly_band.reset_index().to_excel(writer, sheet_name=sheet_name, index=False)

            # Separate intensity tabs because MPA/VPA thresholds overlap cadence bands
            for intensity_label, intensity_mask in [
                (mpa_label, (part_df['Steps'] >= mpa) & (part_df['Steps'] < vpa)),
                (vpa_label, part_df['Steps'] >= vpa)
            ]:
                intensity_data = part_df[intensity_mask]
                hourly_intensity = intensity_data.groupby(
                    ['Date', 'Hour']
                ).size().unstack(fill_value=0)
                hourly_intensity = hourly_intensity.reindex(index=all_dates, fill_value=0)
                hourly_intensity = hourly_intensity.reindex(columns=range(24), fill_value=0)
                hourly_intensity = hourly_intensity.astype(int)
                hourly_intensity.columns = hour_labels
                hourly_intensity.reset_index().to_excel(
                    writer,
                    sheet_name=intensity_label[:31],
                    index=False
                )

        excel_buffer.seek(0)
        red_flag_buffer = build_red_flag_workbook(part_df, manual_participant_id)

        # 6. DISPLAY RESULTS AND DOWNLOAD BUTTONS
        st.success("✅ Analysis Complete!")
        st.caption(
            f"Source timestamps were interpreted as UTC and converted to {timezone} "
            "before aggregation. Date-specific daylight-saving offsets were applied automatically. "
            f"All outputs include {analysis_start_date:%B %d, %Y} through "
            f"{analysis_end_date:%B %d, %Y}."
        )
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            csv_summary = final_summary.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📥 Download Daily Summaries (CSV)",
                data=csv_summary,
                file_name=f"{manual_participant_id}_Daily_Summaries.csv",
                mime="text/csv"
            )
        with col2:
            csv_min_by_min = final_min_by_min.to_csv(index=False).encode('utf-8')
            st.download_button(
                "📥 Download Minute-by-Minute Log (CSV)",
                data=csv_min_by_min,
                file_name=f"{manual_participant_id}_Min_by_Min.csv",
                mime="text/csv"
            )
        with col3:
            st.download_button(
                "📊 Download Hourly Analysis (Excel)",
                data=excel_buffer,
                file_name=f"{manual_participant_id}_Hourly_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col4:
            st.download_button(
                "🚩 Download Red Flag Summary (Excel)",
                data=red_flag_buffer,
                file_name=f"{manual_participant_id}_Red_Flag_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    else:
        st.error("❌ No valid Fitbit data found in the uploaded files. Please check your files and try again.")
